"""Spreadsheet import: all-or-nothing, non-destructive, dry-run capable.

- Accepts the two-sheet .xlsx workbook or a single-sheet .csv (identified by
  its header row); the format is detected from the file content.
- Volunteers are matched by email (plus name to break family-shared-email
  ties). This is NOT a fallback chain: a row carrying an email that matches
  nobody does not then try the name, it creates a volunteer (with a warning if
  that name already exists). Only a row with a blank email matches by name.
- Memberships are upserted; rows never delete existing memberships.
- Any error rolls the whole import back; the report lists every issue.
"""

import base64
import binascii
import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO

import sqlalchemy as sa
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from ..db import db_session
from ..log import audit_log
from ..models import AppUser, Membership, Volunteer
from ..permissions import Actor, load_actor
from ..services import memberships as membership_service
from ..services import photos as photo_service
from ..services import teams as team_service
from .common import (
    FORMULA_STARTERS,
    MEMBERSHIP_HEADERS,
    MEMBERSHIP_SHEET,
    PHOTO_HEADER,
    VOLUNTEER_HEADERS,
    VOLUNTEER_SHEET,
    parse_role,
)


@dataclass
class Issue:
    sheet: str
    row: int
    message: str


@dataclass
class ImportReport:
    volunteers_created: int = 0
    volunteers_updated: int = 0
    memberships_created: int = 0
    memberships_updated: int = 0
    photos_set: int = 0
    errors: list[Issue] = field(default_factory=list)
    warnings: list[Issue] = field(default_factory=list)
    applied: bool = False

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


class _Abort(Exception):
    pass


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if len(text) > 1 and text[0] == "'" and text[1] in FORMULA_STARTERS:
        text = text[1:]  # undo the exporter's formula-injection escape
    return text or None


def _parse_date(value, report: ImportReport, sheet: str, row: int) -> date | None:
    if value is None or _clean(value) is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        report.warnings.append(Issue(sheet, row, f"unreadable date {value!r}, ignored"))
        return None


@dataclass
class _Parsed:
    """Sheet data rows regardless of source format; None means sheet absent."""

    volunteer_rows: list[tuple] | None
    membership_rows: list[tuple] | None
    # the optional Photo column (position 7) was present in the header
    volunteer_has_photo: bool = False


# Accepted spellings of "yes" in the Active column. An empty cell means active;
# anything outside this set archives the volunteer, so unrecognised values are
# warned about rather than applied silently.
ACTIVE_VALUES = ("yes", "y", "true", "1", "x")

_EXTRA_COLUMNS_WARNING = (
    "extra columns (custom fields) are ignored — custom values are not imported yet"
)


def _photo_column_present(header) -> bool:
    """Is the optional Photo column in position 7, right after the base headers?"""
    n = len(VOLUNTEER_HEADERS)
    if len(header) <= n or header[n] is None:
        return False
    return str(header[n]).strip().casefold() == PHOTO_HEADER.casefold()


def _parse_xlsx(content: bytes, report: ImportReport) -> _Parsed | None:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        report.errors.append(Issue("-", 0, f"cannot read workbook: {exc}"))
        return None

    volunteer_rows = membership_rows = None
    has_photo = False
    if VOLUNTEER_SHEET in workbook.sheetnames:
        sheet = workbook[VOLUNTEER_SHEET]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        has_photo = _photo_column_present(header)
        expected = len(VOLUNTEER_HEADERS) + (1 if has_photo else 0)
        if sum(1 for cell in header if cell is not None) > expected:
            report.warnings.append(Issue(VOLUNTEER_SHEET, 1, _EXTRA_COLUMNS_WARNING))
        volunteer_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if MEMBERSHIP_SHEET in workbook.sheetnames:
        membership_rows = list(
            workbook[MEMBERSHIP_SHEET].iter_rows(min_row=2, values_only=True)
        )
    return _Parsed(volunteer_rows, membership_rows, volunteer_has_photo=has_photo)


def _parse_csv(content: bytes, report: ImportReport) -> _Parsed | None:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        report.errors.append(
            Issue("-", 0, "cannot read file: not an .xlsx workbook or a UTF-8 .csv")
        )
        return None

    rows = list(csv.reader(StringIO(text)))
    header = [str(cell).strip().casefold() for cell in rows[0]] if rows else []
    data = [tuple(row) for row in rows[1:]]
    n = len(VOLUNTEER_HEADERS)
    if header[:n] == [h.casefold() for h in VOLUNTEER_HEADERS]:
        # the Photo column is optional so pre-photo 6-column files keep working
        has_photo = _photo_column_present(header)
        if any(header[n + 1 if has_photo else n :]):
            report.warnings.append(Issue(VOLUNTEER_SHEET, 1, _EXTRA_COLUMNS_WARNING))
        return _Parsed(data, None, volunteer_has_photo=has_photo)
    if header[:n] == [h.casefold() for h in MEMBERSHIP_HEADERS]:
        if any(header[n:]):
            report.warnings.append(
                Issue(MEMBERSHIP_SHEET, 1, "extra columns are ignored")
            )
        return _Parsed(None, data)
    report.errors.append(
        Issue(
            "-",
            1,
            "cannot identify CSV: header row matches neither the Volunteers "
            "nor the Memberships sheet",
        )
    )
    return None


async def run_import(
    content: bytes, *, dry_run: bool, user_id: int | None
) -> ImportReport:
    """Parse and apply a workbook or CSV. On dry_run or any error, everything
    rolls back. Non-admin users (leaders/seconds) are scoped to the teams they
    manage; user_id=None runs unrestricted (service-level callers)."""
    report = ImportReport()
    if content[:4] == b"PK\x03\x04":  # zip magic: .xlsx
        parsed = _parse_xlsx(content, report)
    else:
        parsed = _parse_csv(content, report)
    if parsed is None:
        return report

    try:
        async with db_session(user_id) as session:
            actor = None
            if user_id is not None:
                user = await session.get(AppUser, user_id)
                assert user is not None, f"unknown user {user_id}"
                actor = await load_actor(session, user)
            await _apply(session, parsed, report, actor)
            if dry_run or report.has_errors:
                raise _Abort()
            report.applied = True
    except _Abort:
        pass
    audit_log(
        "import.finished",
        outcome=(
            "applied"
            if report.applied
            else "dry-run (rolled back)"
            if dry_run
            else "failed (rolled back)"
        ),
        volunteers_created=report.volunteers_created,
        volunteers_updated=report.volunteers_updated,
        memberships_created=report.memberships_created,
        memberships_updated=report.memberships_updated,
        photos_set=report.photos_set,
        errors=len(report.errors),
    )
    return report


async def _apply_photo(
    session: AsyncSession,
    volunteer: Volunteer,
    photo_raw: str,
    report: ImportReport,
    row_num: int,
    actor: Actor | None,
) -> None:
    """Decode and store one Photo cell; row-level errors, never an exception."""
    try:
        decoded = base64.b64decode(photo_raw, validate=True)
    except (binascii.Error, ValueError):
        report.errors.append(
            Issue(VOLUNTEER_SHEET, row_num, "Photo is not valid base64")
        )
        return
    if len(decoded) > photo_service.MAX_UPLOAD_BYTES:
        report.errors.append(
            Issue(VOLUNTEER_SHEET, row_num, "Photo is larger than 10 MB")
        )
        return
    # byte-equal to what's stored (the usual export -> import round-trip):
    # skip, so re-imports are no-ops instead of JPEG re-compression churn
    existing = await photo_service.get(session, volunteer.id)
    if existing is not None and existing.image == decoded:
        return
    try:
        await photo_service.set_photo(
            session,
            volunteer.id,
            decoded,
            uploaded_by=actor.user.id if actor is not None else None,
        )
    except ValueError as exc:
        report.errors.append(Issue(VOLUNTEER_SHEET, row_num, f"Photo: {exc}"))
        return
    report.photos_set += 1


async def _apply(
    session: AsyncSession,
    parsed: _Parsed,
    report: ImportReport,
    actor: Actor | None = None,
) -> None:
    volunteers = list((await session.execute(sa.select(Volunteer))).scalars())
    by_email: dict[str, list[Volunteer]] = {}
    by_name: dict[str, list[Volunteer]] = {}

    def index(v: Volunteer) -> None:
        if v.email:
            by_email.setdefault(v.email.lower(), []).append(v)
        by_name.setdefault(v.full_name.lower(), []).append(v)

    for v in volunteers:
        index(v)

    def match(email: str | None, name: str | None) -> Volunteer | None | str:
        """A volunteer, None (no match), or an error message string."""
        candidates: list[Volunteer] = []
        if email:
            candidates = by_email.get(email.lower(), [])
            if len(candidates) > 1 and name:
                narrowed = [
                    c for c in candidates if c.full_name.lower() == name.lower()
                ]
                if narrowed:
                    candidates = narrowed
        elif name:
            candidates = by_name.get(name.lower(), [])
        if len(candidates) > 1:
            return f"ambiguous: {len(candidates)} volunteers match {email or name!r}"
        return candidates[0] if candidates else None

    restricted = actor is not None and not actor.is_admin

    # every current membership in one query: the upsert loop and the restricted
    # scope check both read from this map instead of issuing per-row lookups
    membership_by_pair: dict[tuple[int, int], Membership] = {
        (m.volunteer_id, m.team_id): m
        for m in (await session.execute(sa.select(Membership))).scalars()
    }

    all_teams = await team_service.list_all(session)
    paths = team_service.team_paths(all_teams)
    team_by_path = {p.lower(): tid for tid, p in paths.items()}
    team_by_name: dict[str, list[int]] = {}
    for t in all_teams:
        team_by_name.setdefault(t.name.lower(), []).append(t.id)

    def resolve_team(team_path: str) -> int | str:
        """A team id, or an error message string."""
        team_id = team_by_path.get(team_path.lower())
        if team_id is not None:
            return team_id
        ids = team_by_name.get(team_path.lower(), [])
        if len(ids) == 1:
            return ids[0]
        if len(ids) > 1:
            return f"team name {team_path!r} is ambiguous, use its full path"
        return f"unknown team {team_path!r}"

    # Restricted imports (leaders/seconds): a Memberships row that puts someone
    # on a managed team licenses that person's Volunteers row too — an existing
    # volunteer's contact update (granted_ids, keyed by resolved id so a
    # family-shared email cannot license the other spouse), or a brand-new
    # volunteer's creation (granted_new_keys). Scope is pre-import state only.
    granted_ids: set[int] = set()
    granted_new_keys: set[str] = set()
    teams_by_volunteer: dict[int, set[int]] = {}
    if restricted:
        for volunteer_id, team_id in membership_by_pair:
            teams_by_volunteer.setdefault(volunteer_id, set()).add(team_id)
        for row in parsed.membership_rows or []:
            row = tuple(row) + (None,) * (3 - len(row))
            email, name, team_path = (_clean(c) for c in row[:3])
            if not team_path:
                continue
            team_id = resolve_team(team_path)
            if isinstance(team_id, str) or team_id not in actor.managed_team_ids:
                continue
            found = match(email, name)
            if isinstance(found, str):
                continue
            if found is None:
                if email:
                    granted_new_keys.add(email.lower())
                if name:
                    granted_new_keys.add(name.lower())
            else:
                granted_ids.add(found.id)

    # --- Volunteers sheet ---
    if parsed.volunteer_rows is not None:
        for row_num, row in enumerate(parsed.volunteer_rows, start=2):
            row = tuple(row) + (None,) * (7 - len(row))
            first, last, email, phone, notes, active = (_clean(c) for c in row[:6])
            photo_raw = _clean(row[6]) if parsed.volunteer_has_photo else None
            if not first and not last:
                continue
            if not first or not last:
                report.errors.append(
                    Issue(
                        VOLUNTEER_SHEET,
                        row_num,
                        "first and last name are both required",
                    )
                )
                continue
            email = email.lower() if email else None
            full_name = f"{first} {last}"
            found = match(email, full_name)
            if isinstance(found, str):
                report.errors.append(Issue(VOLUNTEER_SHEET, row_num, found))
                continue
            is_active = active is None or active.lower() in ACTIVE_VALUES
            says_no = active is not None and active.lower() in ("no", "n", "false", "0")
            if active is not None and not is_active and not says_no:
                report.warnings.append(
                    Issue(
                        VOLUNTEER_SHEET,
                        row_num,
                        f"Active value {active!r} is not recognised — archiving "
                        f"{full_name}. Use one of {', '.join(ACTIVE_VALUES)} to keep them active.",
                    )
                )
            if found is None:
                # Matching is email-first with no name fallback, so a new address
                # for someone already on file silently creates a second record.
                # That is the intended rule; saying nothing about it is what made
                # the July import produce duplicates.
                if email and by_name.get(full_name.lower()):
                    report.warnings.append(
                        Issue(
                            VOLUNTEER_SHEET,
                            row_num,
                            f"{email!r} matched nobody, but {full_name!r} already exists — "
                            "creating a NEW volunteer. If they are the same person, set the "
                            "email on the existing record first.",
                        )
                    )
                if restricted and not (
                    (email and email in granted_new_keys)
                    or full_name.lower() in granted_new_keys
                ):
                    report.errors.append(
                        Issue(
                            VOLUNTEER_SHEET,
                            row_num,
                            "new volunteers must also be added to a team you lead "
                            "(Memberships rows)",
                        )
                    )
                    continue
                v = Volunteer(
                    first_name=first,
                    last_name=last,
                    email=email,
                    phone=phone,
                    notes=notes,
                    is_active=is_active,
                )
                session.add(v)
                await session.flush()
                index(v)
                report.volunteers_created += 1
                target = v
            else:
                # Denied even when the row changes nothing: erroring only on a
                # change would let dry-runs probe contact fields by brute force.
                if restricted and not (
                    actor.can_edit_volunteer(
                        found.id, teams_by_volunteer.get(found.id, set())
                    )
                    or found.id in granted_ids
                ):
                    report.errors.append(
                        Issue(
                            VOLUNTEER_SHEET,
                            row_num,
                            f"not allowed to edit volunteer {email or f'{first} {last}'!r} "
                            "(not on a team you lead)",
                        )
                    )
                    continue
                changed = False
                for attr, value in (
                    ("first_name", first),
                    ("last_name", last),
                    ("email", email),
                    ("phone", phone),
                    ("notes", notes),
                    ("is_active", is_active),
                ):
                    if value is not None and getattr(found, attr) != value:
                        setattr(found, attr, value)
                        changed = True
                if changed:
                    report.volunteers_updated += 1
                target = found

            # a blank Photo cell means "leave unchanged", like every other
            # column; removal happens in the app/API, never via spreadsheet
            if photo_raw is not None:
                await _apply_photo(session, target, photo_raw, report, row_num, actor)

    # --- Memberships sheet ---
    if parsed.membership_rows is not None:
        for row_num, row in enumerate(parsed.membership_rows, start=2):
            row = tuple(row) + (None,) * (6 - len(row))
            email, name, team_path, role_raw, joined_raw, notes = row[:6]
            email, name, team_path = _clean(email), _clean(name), _clean(team_path)
            if not any((email, name, team_path, _clean(role_raw))):
                continue
            if not team_path:
                report.errors.append(
                    Issue(MEMBERSHIP_SHEET, row_num, "team path is required")
                )
                continue

            team_id = resolve_team(team_path)
            if isinstance(team_id, str):
                report.errors.append(Issue(MEMBERSHIP_SHEET, row_num, team_id))
                continue
            if restricted and team_id not in actor.managed_team_ids:
                report.errors.append(
                    Issue(
                        MEMBERSHIP_SHEET,
                        row_num,
                        f"not allowed: {team_path!r} is not a team you lead",
                    )
                )
                continue

            role = parse_role(role_raw) if role_raw is not None else None
            if role is None:
                report.errors.append(
                    Issue(MEMBERSHIP_SHEET, row_num, f"unknown role {role_raw!r}")
                )
                continue

            found = match(email, name)
            if isinstance(found, str):
                report.errors.append(Issue(MEMBERSHIP_SHEET, row_num, found))
                continue
            if found is None:
                report.errors.append(
                    Issue(
                        MEMBERSHIP_SHEET,
                        row_num,
                        f"unknown volunteer {email or name!r} (add them to the Volunteers sheet)",
                    )
                )
                continue

            joined_on = _parse_date(joined_raw, report, MEMBERSHIP_SHEET, row_num)
            existing = membership_by_pair.get((found.id, team_id))
            before = (
                (existing.role, existing.joined_on, existing.notes)
                if existing
                else None
            )
            membership = await membership_service.assign(
                session,
                found.id,
                team_id,
                role,
                joined_on=joined_on,
                notes=_clean(notes),
                existing=existing,
            )
            if existing is None:
                # a later sheet row for the same pair must hit the update branch
                membership_by_pair[(found.id, team_id)] = membership
                report.memberships_created += 1
            elif before != (membership.role, membership.joined_on, membership.notes):
                report.memberships_updated += 1
