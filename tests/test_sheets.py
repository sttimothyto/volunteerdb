"""Spreadsheet export → import round-trip, validation, and dry-run."""

from io import BytesIO

from openpyxl import load_workbook

from volunteerdb.db import db_session
from volunteerdb.models import FieldType, TeamRole
from volunteerdb.services import custom_fields, memberships, teams, volunteers
from volunteerdb.sheets import exporter, importer
from volunteerdb.sheets.common import MEMBERSHIP_SHEET, VOLUNTEER_SHEET


async def _setup(session):
    liturgy = await teams.create(session, "Liturgy")
    music = await teams.create(session, "Music", parent_team_id=liturgy.id)
    anna = await volunteers.create(
        session, "Anna", "Smith", "anna@example.org", phone="555-1"
    )
    ben = await volunteers.create(session, "Ben", "Jones", "ben@example.org")
    await memberships.assign(session, anna.id, liturgy.id, TeamRole.leader)
    await memberships.assign(session, ben.id, music.id, TeamRole.member)
    return liturgy, music, anna, ben


async def test_roundtrip_reimport_is_a_noop(database):
    async with db_session() as session:
        await _setup(session)
        content = await exporter.export_workbook(session)

    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors
    assert report.applied
    assert report.volunteers_created == 0
    assert report.volunteers_updated == 0
    assert report.memberships_created == 0
    assert report.memberships_updated == 0


async def test_import_applies_edits_and_additions(database):
    async with db_session() as session:
        await _setup(session)
        content = await exporter.export_workbook(session)

    wb = load_workbook(BytesIO(content))
    vs, ms = wb[VOLUNTEER_SHEET], wb[MEMBERSHIP_SHEET]
    # promote Ben to Music leader (role is column D on his existing row)
    for row in ms.iter_rows(min_row=2):
        if row[0].value == "ben@example.org":
            row[3].value = "Ministry leader"
    # add a brand-new volunteer with a membership by team path
    vs.append(["Cara", "White", "cara@example.org", "555-9", None, "yes"])
    ms.append(
        ["cara@example.org", "Cara White", "Liturgy / Music", "member", None, None]
    )
    buffer = BytesIO()
    wb.save(buffer)

    report = await importer.run_import(buffer.getvalue(), dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert report.applied
    assert report.volunteers_created == 1
    assert report.memberships_created == 1
    assert report.memberships_updated == 1

    async with db_session() as session:
        cara = (await volunteers.search(session, "Cara"))[0]
        assert cara.email == "cara@example.org"
        found = await volunteers.search(session, "Ben")
        ben_assignments = await volunteers.assignments(session, found[0].id)
        assert ben_assignments[0][0].role == TeamRole.leader


async def test_unknown_team_blocks_everything(database):
    async with db_session() as session:
        await _setup(session)
        content = await exporter.export_workbook(session)

    wb = load_workbook(BytesIO(content))
    wb[VOLUNTEER_SHEET].append(["Dave", "Black", "dave@example.org", None, None, "yes"])
    wb[MEMBERSHIP_SHEET].append(
        ["dave@example.org", "Dave Black", "No Such Team", "member", None, None]
    )
    buffer = BytesIO()
    wb.save(buffer)

    report = await importer.run_import(buffer.getvalue(), dry_run=False, user_id=None)
    assert report.has_errors
    assert not report.applied
    assert "No Such Team" in report.errors[0].message

    async with db_session() as session:
        assert await volunteers.search(session, "Dave") == [], (
            "all-or-nothing: Dave not created"
        )


async def test_dry_run_writes_nothing(database):
    async with db_session() as session:
        await teams.create(session, "Liturgy")

    wb = load_workbook(BytesIO(exporter.template_workbook()))
    wb[VOLUNTEER_SHEET].append(["Eve", "Green", "eve@example.org", None, None, "yes"])
    wb[MEMBERSHIP_SHEET].append(
        ["eve@example.org", "Eve Green", "Liturgy", "leader", None, None]
    )
    buffer = BytesIO()
    wb.save(buffer)

    report = await importer.run_import(buffer.getvalue(), dry_run=True, user_id=None)
    assert not report.has_errors
    assert not report.applied
    assert report.volunteers_created == 1  # would be created

    async with db_session() as session:
        assert await volunteers.search(session, "Eve") == []


async def test_export_includes_custom_columns_and_reimport_ignores_them(database):
    async with db_session() as session:
        _, _, anna, _ = await _setup(session)
        await custom_fields.create_def(session, "Shirt size", FieldType.text)
        await custom_fields.create_def(session, "Trained", FieldType.checkbox)
        await custom_fields.set_values(
            session, anna.id, {"shirt_size": "M", "trained": True}
        )
        content = await exporter.export_workbook(session)

    vs = load_workbook(BytesIO(content))[VOLUNTEER_SHEET]
    headers = [c.value for c in vs[1]]
    assert headers[-2:] == ["Shirt size", "Trained"]
    anna_row = next(
        r for r in vs.iter_rows(min_row=2, values_only=True) if r[0] == "Anna"
    )
    # custom values sit after the Photo column (index 6)
    assert anna_row[7] == "M" and anna_row[8] == "yes"

    # round-trip stays safe: the extra columns are ignored with a warning
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors
    assert report.applied
    assert report.volunteers_updated == 0
    assert any("custom" in w.message for w in report.warnings)

    async with db_session() as session:
        (found,) = await volunteers.search(session, "Anna")
        assert found.custom == {"shirt_size": "M", "trained": True}, "values untouched"
