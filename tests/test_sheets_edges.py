"""Spreadsheet edge cases: injection escapes, date/role parsing, row validation."""

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from volunteerdb.db import db_session
from volunteerdb.models import TeamRole
from volunteerdb.services import teams, volunteers
from volunteerdb.sheets import exporter, importer
from volunteerdb.sheets.common import MEMBERSHIP_SHEET, VOLUNTEER_SHEET, parse_role
from volunteerdb.sheets.exporter import _safe
from volunteerdb.sheets.importer import ImportReport, _clean, _parse_date


def _workbook_bytes(volunteer_rows: list = (), membership_rows: list = ()) -> bytes:
    wb = load_workbook(BytesIO(exporter.template_workbook()))
    for row in volunteer_rows:
        wb[VOLUNTEER_SHEET].append(row)
    for row in membership_rows:
        wb[MEMBERSHIP_SHEET].append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_clean_and_safe_units():
    assert _safe("=SUM(A1)") == "'=SUM(A1)"
    assert _safe("plain text") == "plain text"
    assert _safe(None) is None
    assert _safe(5) == 5, "non-strings pass through untouched"

    assert _clean("'=SUM(A1)") == "=SUM(A1)"
    assert _clean("  padded  ") == "padded"
    assert _clean("   ") is None
    assert _clean(None) is None
    assert _clean(_safe("=1+1")) == "=1+1", "escape/unescape are inverses"


def test_parse_date_variants():
    report = ImportReport()
    assert _parse_date(datetime(2024, 3, 5, 10, 30), report, "S", 2) == date(2024, 3, 5)
    assert _parse_date(date(2024, 3, 5), report, "S", 3) == date(2024, 3, 5)
    assert _parse_date("2024-03-05", report, "S", 4) == date(2024, 3, 5)
    assert _parse_date(None, report, "S", 5) is None
    assert _parse_date("   ", report, "S", 6) is None
    assert report.warnings == [], "no warnings so far"

    assert _parse_date("05/03/2024", report, "S", 7) is None
    (warning,) = report.warnings
    assert warning.row == 7 and "unreadable date" in warning.message


def test_parse_role_value_and_label():
    assert parse_role("leader") == TeamRole.leader
    assert parse_role("Ministry leader") == TeamRole.leader
    assert parse_role("MEMBER") == TeamRole.member
    assert parse_role("Core team member") == TeamRole.core
    assert parse_role(" second ") == TeamRole.second
    assert parse_role("bogus") is None


async def test_formula_injection_escape_roundtrips(database):
    async with db_session() as session:
        await volunteers.create(
            session, "=Evil", "Person", "evil@example.org", notes="=SUM(A1:A9)"
        )
        content = await exporter.export_workbook(session)

    vs = load_workbook(BytesIO(content))[VOLUNTEER_SHEET]
    row = next(r for r in vs.iter_rows(min_row=2, values_only=True) if r[2] == "evil@example.org")
    assert row[0] == "'=Evil" and row[4] == "'=SUM(A1:A9)", "cells never start with a bare ="

    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors
    assert report.volunteers_updated == 0, "round-trip is a no-op"

    async with db_session() as session:
        (found,) = await volunteers.search(session, "evil@example.org")
        assert found.first_name == "=Evil" and found.notes == "=SUM(A1:A9)", (
            "database keeps the raw values"
        )


async def test_import_row_validation_errors(database):
    async with db_session() as session:
        await teams.create(session, "Liturgy")

    content = _workbook_bytes(
        volunteer_rows=[["OnlyFirst", None, None, None, None, "yes"]],
        membership_rows=[
            ["ghost@example.org", "No Body", "Liturgy", "member", None, None],
            [None, None, None, "member", None, None],
            ["ghost@example.org", "No Body", "Liturgy", "grand-poobah", None, None],
        ],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)

    assert not report.applied, "all-or-nothing"
    messages = [issue.message for issue in report.errors]
    assert any("first and last name are both required" in m for m in messages)
    assert any("unknown volunteer" in m for m in messages)
    assert any("team path is required" in m for m in messages)
    assert any("unknown role 'grand-poobah'" in m for m in messages)


async def test_import_ambiguous_matches_error(database):
    async with db_session() as session:
        liturgy = await teams.create(session, "Liturgy")
        youth = await teams.create(session, "Youth")
        await teams.create(session, "Music", parent_team_id=liturgy.id)
        await teams.create(session, "Music", parent_team_id=youth.id)
        await volunteers.create(session, "Sam", "Same")
        await volunteers.create(session, "Sam", "Same")
        await volunteers.create(session, "Uma", "Unique", "uma@example.org")

    content = _workbook_bytes(
        membership_rows=[
            [None, "Sam Same", "Liturgy", "member", None, None],
            ["uma@example.org", "Uma Unique", "Music", "member", None, None],
        ],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)

    assert not report.applied
    messages = [issue.message for issue in report.errors]
    assert any("ambiguous: 2 volunteers match 'Sam Same'" in m for m in messages)
    assert any("'Music' is ambiguous, use its full path" in m for m in messages)


async def test_active_column_blank_counts_as_active(database):
    content = _workbook_bytes(
        volunteer_rows=[["Blank", "Active", "blank@example.org", None, None, None]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert report.applied and report.volunteers_created == 1

    async with db_session() as session:
        (found,) = await volunteers.search(session, "blank@example.org")
        assert found.is_active is True, "an empty Active cell means active"
