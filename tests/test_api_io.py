"""Spreadsheet export/import over HTTP: permissions, size cap, dry-run."""

from io import BytesIO

from openpyxl import load_workbook

from volunteerdb.db import db_session
from volunteerdb.models import TeamRole
from volunteerdb.services import memberships, teams, volunteers
from volunteerdb.sheets import exporter
from volunteerdb.sheets.common import (
    MEMBERSHIP_HEADERS,
    MEMBERSHIP_SHEET,
    VOLUNTEER_HEADERS,
    VOLUNTEER_SHEET,
)

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _upload(content: bytes) -> dict:
    return {"file": ("import.xlsx", content, XLSX_TYPE)}


def _upload_csv(content: bytes) -> dict:
    return {"file": ("import.csv", content, "text/csv")}


def _import_workbook_bytes(rows_volunteers: list, rows_memberships: list) -> bytes:
    wb = load_workbook(BytesIO(exporter.template_workbook()))
    for row in rows_volunteers:
        wb[VOLUNTEER_SHEET].append(row)
    for row in rows_memberships:
        wb[MEMBERSHIP_SHEET].append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def test_template_download(client, seeded, token_member):
    r = await client.get("/api/export/template.xlsx", headers=token_member)
    assert r.status_code == 200, "any signed-in user may fetch the template"
    assert r.headers["content-type"] == XLSX_TYPE
    assert 'filename="volunteerdb-template.xlsx"' in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert set(wb.sheetnames) == {VOLUNTEER_SHEET, MEMBERSHIP_SHEET}
    assert [c.value for c in wb[VOLUNTEER_SHEET][1]] == VOLUNTEER_HEADERS
    assert [c.value for c in wb[MEMBERSHIP_SHEET][1]] == MEMBERSHIP_HEADERS


async def test_parish_export_admin_only(client, seeded, token_admin, token_member):
    r = await client.get("/api/export/parish.xlsx", headers=token_member)
    assert r.status_code == 403

    r = await client.get("/api/export/parish.xlsx", headers=token_admin)
    assert r.status_code == 200
    vs = load_workbook(BytesIO(r.content))[VOLUNTEER_SHEET]
    names = {(row[0], row[1]) for row in vs.iter_rows(min_row=2, values_only=True)}
    assert ("Maria", "Alvarez") in names


async def test_team_export_permission_matrix(client, seeded, token_admin, token_member):
    team_id = seeded["team_id"]

    r = await client.get(f"/api/export/team/{team_id}.xlsx", headers=token_member)
    assert r.status_code == 403, "a plain member cannot export contact details"

    async with db_session() as session:
        # promote the member's volunteer to core: full-roster rights incl. sub-teams
        await memberships.assign(session, seeded["volunteer_id"], team_id, TeamRole.core)
        music = await teams.create(session, "Music", parent_team_id=team_id)
        singer = await volunteers.create(session, "Sally", "Singer", "sally@example.org")
        await memberships.assign(session, singer.id, music.id, TeamRole.member)

    r = await client.get(f"/api/export/team/{team_id}.xlsx", headers=token_member)
    assert r.status_code == 200
    ms = load_workbook(BytesIO(r.content))[MEMBERSHIP_SHEET]
    paths = {row[2] for row in ms.iter_rows(min_row=2, values_only=True)}
    assert paths == {"Liturgy", "Liturgy / Music"}, "subtree included"

    r = await client.get(f"/api/export/team/{team_id}.xlsx", headers=token_admin)
    assert r.status_code == 200


async def test_csv_template_download(client, seeded, token_member):
    r = await client.get("/api/export/template/volunteers.csv", headers=token_member)
    assert r.status_code == 200, "any signed-in user may fetch the template"
    assert r.headers["content-type"].startswith("text/csv")
    assert r.content.startswith(b"\xef\xbb\xbf"), "UTF-8 BOM for Excel"
    assert r.content.decode("utf-8-sig").splitlines()[0] == ",".join(VOLUNTEER_HEADERS)

    r = await client.get("/api/export/template/bogus.csv", headers=token_member)
    assert r.status_code == 422


async def test_parish_csv_admin_only(client, seeded, token_admin, token_member):
    r = await client.get("/api/export/parish/volunteers.csv", headers=token_member)
    assert r.status_code == 403

    r = await client.get("/api/export/parish/volunteers.csv", headers=token_admin)
    assert r.status_code == 200
    assert "Maria" in r.content.decode("utf-8-sig")
    assert 'filename="volunteerdb-parish-volunteers.csv"' in r.headers["content-disposition"]


async def test_team_csv_permission_mirrors_xlsx(client, seeded, token_admin, token_member):
    team_id = seeded["team_id"]
    r = await client.get(f"/api/export/team/{team_id}/memberships.csv", headers=token_member)
    assert r.status_code == 403, "a plain member cannot export contact details"

    r = await client.get(f"/api/export/team/{team_id}/memberships.csv", headers=token_admin)
    assert r.status_code == 200
    assert "Liturgy" in r.content.decode("utf-8-sig")


async def test_my_teams_export(client, seeded, token_leader, token_member):
    for path in ("/api/export/my-teams.xlsx", "/api/export/my-teams/volunteers.csv"):
        r = await client.get(path, headers=token_member)
        assert r.status_code == 403, "no managed teams, no scoped export"

    r = await client.get("/api/export/my-teams.xlsx", headers=token_leader)
    assert r.status_code == 200
    ms = load_workbook(BytesIO(r.content))[MEMBERSHIP_SHEET]
    paths = {row[2] for row in ms.iter_rows(min_row=2, values_only=True)}
    assert paths == {"Liturgy"}

    r = await client.get("/api/export/my-teams/volunteers.csv", headers=token_leader)
    assert r.status_code == 200
    body = r.content.decode("utf-8-sig")
    assert "Maria" in body and "Lena" in body


async def test_leader_import_scoped_over_http(client, seeded, token_leader):
    # in scope: update Maria's phone and promote her on Liturgy
    content = _import_workbook_bytes(
        [["Maria", "Alvarez", "maria@example.org", "555-42", None, "yes"]],
        [["maria@example.org", "Maria Alvarez", "Liturgy", "core", None, None]],
    )
    r = await client.post("/api/import", files=_upload(content), headers=token_leader)
    assert r.status_code == 200, r.text
    report = r.json()
    assert report["applied"] is True
    assert report["volunteers_updated"] == 1 and report["memberships_updated"] == 1

    # out of scope: a team the leader does not manage → row error, nothing applied
    async with db_session() as session:
        await teams.create(session, "Hospitality")
    content = _import_workbook_bytes(
        [], [["maria@example.org", "Maria Alvarez", "Hospitality", "member", None, None]]
    )
    r = await client.post("/api/import", files=_upload(content), headers=token_leader)
    assert r.status_code == 200
    report = r.json()
    assert report["applied"] is False
    assert any("not a team you lead" in e["message"] for e in report["errors"])


async def test_leader_csv_dry_run_over_http(client, seeded, token_leader):
    content = (
        ",".join(VOLUNTEER_HEADERS) + "\nMaria,Alvarez,maria@example.org,555-77,,yes\n"
    ).encode("utf-8-sig")
    r = await client.post(
        "/api/import", params={"dry_run": "true"}, files=_upload_csv(content), headers=token_leader
    )
    assert r.status_code == 200, r.text
    report = r.json()
    assert report["applied"] is False and report["volunteers_updated"] == 1
    assert not report["errors"]


async def test_import_dry_run_then_apply_over_http(client, seeded, token_admin, token_member):
    content = _import_workbook_bytes(
        [["Eve", "Green", "eve@example.org", None, None, "yes"]],
        [["eve@example.org", "Eve Green", "Liturgy", "leader", None, None]],
    )

    r = await client.post("/api/import", files=_upload(content), headers=token_member)
    assert r.status_code == 403

    r = await client.post(
        "/api/import", params={"dry_run": "true"}, files=_upload(content), headers=token_admin
    )
    assert r.status_code == 200
    report = r.json()
    assert report["applied"] is False
    assert report["volunteers_created"] == 1 and report["memberships_created"] == 1
    async with db_session() as session:
        assert await volunteers.search(session, "Eve") == [], "dry run wrote nothing"

    r = await client.post("/api/import", files=_upload(content), headers=token_admin)
    assert r.status_code == 200
    assert r.json()["applied"] is True
    async with db_session() as session:
        (eve,) = await volunteers.search(session, "Eve")
        assert eve.email == "eve@example.org"


async def test_import_oversize_rejected_413(client, seeded, token_admin):
    r = await client.post(
        "/api/import", files=_upload(b"x" * 10_000_001), headers=token_admin
    )
    assert r.status_code == 413


async def test_import_unreadable_file_reports_error(client, seeded, token_admin):
    # not a zip → treated as CSV; unrecognizable header row
    r = await client.post(
        "/api/import", files=_upload(b"this is not a workbook"), headers=token_admin
    )
    assert r.status_code == 200
    report = r.json()
    assert report["applied"] is False
    assert report["errors"] and "cannot identify CSV" in report["errors"][0]["message"]

    # a corrupt zip is still reported as an unreadable workbook
    r = await client.post(
        "/api/import", files=_upload(b"PK\x03\x04 corrupt"), headers=token_admin
    )
    report = r.json()
    assert report["errors"] and "cannot read workbook" in report["errors"][0]["message"]
