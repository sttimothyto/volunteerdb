"""Photo column round-trip: base64 in both xlsx and CSV, optional on import."""

import base64
import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook
from PIL import Image

from volunteerdb.db import db_session
from volunteerdb.services import photos, volunteers
from volunteerdb.sheets import exporter, importer
from volunteerdb.sheets.common import PHOTO_HEADER, VOLUNTEER_HEADERS, VOLUNTEER_SHEET


def _png(width: int = 450, height: int = 450) -> bytes:
    buffer = BytesIO()
    Image.new("RGB", (width, height), (120, 90, 60)).save(buffer, format="PNG")
    return buffer.getvalue()


def _csv_bytes(header, rows) -> bytes:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


async def _seed_with_photo():
    async with db_session() as session:
        anna = await volunteers.create(session, "Anna", "Smith", "anna@example.org")
        stored = await photos.set_photo(session, anna.id, _png(), uploaded_by=None)
        return anna.id, bytes(stored.image)


async def test_xlsx_export_carries_base64_and_reimports_as_noop(database):
    anna_id, stored = await _seed_with_photo()
    async with db_session() as session:
        content = await exporter.export_workbook(session)

    vs = load_workbook(BytesIO(content))[VOLUNTEER_SHEET]
    assert [c.value for c in vs[1]] == [*VOLUNTEER_HEADERS, PHOTO_HEADER]
    anna_row = next(
        r for r in vs.iter_rows(min_row=2, values_only=True) if r[0] == "Anna"
    )
    cell = anna_row[6]
    assert cell and len(cell) <= 32_767, "base64 must fit an Excel cell"
    assert base64.b64decode(cell, validate=True) == stored

    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert report.applied and report.photos_set == 0, "byte-identical photo is skipped"


async def test_csv_export_reimports_photo_as_noop(database):
    anna_id, stored = await _seed_with_photo()
    async with db_session() as session:
        content = await exporter.export_csv(session, "volunteers")

    rows = list(csv.reader(StringIO(content.decode("utf-8-sig"))))
    assert rows[0] == [*VOLUNTEER_HEADERS, PHOTO_HEADER]
    assert base64.b64decode(rows[1][6], validate=True) == stored

    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert report.photos_set == 0


async def test_csv_import_sets_photo_on_new_and_existing_volunteers(database):
    async with db_session() as session:
        anna = await volunteers.create(session, "Anna", "Smith", "anna@example.org")
        anna_id = anna.id

    photo_b64 = base64.b64encode(_png()).decode("ascii")
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER],
        [
            ["Anna", "Smith", "anna@example.org", "", "", "yes", photo_b64],
            ["Noah", "New", "noah@example.org", "", "", "yes", photo_b64],
        ],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert report.applied and report.volunteers_created == 1 and report.photos_set == 2

    async with db_session() as session:
        stored = await photos.get(session, anna_id)
        assert stored is not None
        image = Image.open(BytesIO(stored.image))
        assert image.size == (photos.PHOTO_SIZE, photos.PHOTO_SIZE), (
            "import normalizes too"
        )
        (noah,) = await volunteers.search(session, "Noah")
        assert await photos.get(session, noah.id) is not None, (
            "photo lands on created rows"
        )


async def test_blank_photo_cell_leaves_photo_unchanged(database):
    anna_id, stored = await _seed_with_photo()
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER],
        [["Anna", "Smith", "anna@example.org", "", "", "yes", ""]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors and report.photos_set == 0
    async with db_session() as session:
        assert (await photos.get(session, anna_id)).image == stored, (
            "blank never clears — removal happens in the app, not via spreadsheet"
        )


async def test_bad_base64_is_a_row_error_and_blocks_the_import(database):
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER],
        [["Bad", "Cell", "bad@example.org", "", "", "yes", "!!! not base64 !!!"]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert report.has_errors and not report.applied
    assert any("not valid base64" in e.message for e in report.errors)
    async with db_session() as session:
        assert await volunteers.search(session, "Bad") == [], "all-or-nothing"


async def test_undecodable_image_is_a_row_error(database):
    garbage_b64 = base64.b64encode(b"valid base64, not an image").decode("ascii")
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER],
        [["Gia", "Garble", "gia@example.org", "", "", "yes", garbage_b64]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert report.has_errors and not report.applied
    assert any("Photo:" in e.message for e in report.errors)


async def test_dry_run_counts_photos_but_writes_nothing(database):
    async with db_session() as session:
        anna = await volunteers.create(session, "Anna", "Smith", "anna@example.org")
        anna_id = anna.id
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER],
        [
            [
                "Anna",
                "Smith",
                "anna@example.org",
                "",
                "",
                "yes",
                base64.b64encode(_png()).decode(),
            ]
        ],
    )
    report = await importer.run_import(content, dry_run=True, user_id=None)
    assert not report.has_errors and not report.applied and report.photos_set == 1
    async with db_session() as session:
        assert await photos.get(session, anna_id) is None


async def test_legacy_six_column_csv_still_imports(database):
    """The compat guarantee: pre-photo exports keep working, photos untouched."""
    anna_id, stored = await _seed_with_photo()
    content = _csv_bytes(
        VOLUNTEER_HEADERS,
        [["Anna", "Smith", "anna@example.org", "555-2", "", "yes"]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert report.applied and report.volunteers_updated == 1 and report.photos_set == 0
    async with db_session() as session:
        assert (await photos.get(session, anna_id)).image == stored


async def test_photo_header_with_trailing_custom_columns_still_warns(database):
    async with db_session() as session:
        await volunteers.create(session, "Anna", "Smith", "anna@example.org")
    content = _csv_bytes(
        [*VOLUNTEER_HEADERS, PHOTO_HEADER, "Shirt size"],
        [["Anna", "Smith", "anna@example.org", "", "", "yes", "", "M"]],
    )
    report = await importer.run_import(content, dry_run=False, user_id=None)
    assert not report.has_errors, report.errors
    assert any("custom" in w.message for w in report.warnings)
